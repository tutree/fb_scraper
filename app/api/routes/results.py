import io
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import cast, String
from typing import Optional, List
from uuid import UUID
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ...core.database import get_db
from ...core.logging_config import get_logger
from ...services.scraper import ScraperService
from ...services.gemini_classifier import GeminiClassifier
from ...services.enformion_service import EnformionService
from ...schemas.search_result import (
    SearchResultResponse,
    SearchResultList,
    RecentProcessedItem,
    RecentProcessedList,
    SearchResultUpdate,
    AnalyzeBatchRequest,
    AnalyzeBatchResponse,
    AnalyzeSingleResponse,
    AnalyzeResultItem,
    GeoClassificationOut,
    EnrichResultItem,
    EnrichSingleResponse,
    EnrichBatchRequest,
    EnrichBatchResponse,
)
from ...schemas.post_comment import PostCommentResponse
from ...models.search_result import SearchResult, ResultStatus, UserType
from ...models.post_comment import PostComment
from ...utils.validators import (
    clean_facebook_location,
    clean_facebook_name,
    coerce_is_us_boolean,
    parse_facebook_date,
    is_enrichable,
)
from ...services.classification_prompts import should_remove_not_tutoring_related
from ...services.search_result_cleanup import delete_search_result_and_comments

router = APIRouter(prefix="/results", tags=["results"])

logger = get_logger(__name__)

_NOT_ARCHIVED = SearchResult.archived.is_(False)


def _geo_raw_to_out(g: dict) -> GeoClassificationOut:
    return GeoClassificationOut(
        is_us=coerce_is_us_boolean(g.get("is_us")),
        confidence=float(g.get("confidence", 0.0)),
        reason=str(g.get("reason", "")),
    )


def _format_analyze_item(
    result: SearchResult,
    success: bool,
    message: str,
    geo: Optional[GeoClassificationOut] = None,
) -> AnalyzeResultItem:
    return AnalyzeResultItem(
        id=result.id,
        success=success,
        message=message,
        user_type=result.user_type.value if result.user_type else None,
        confidence_score=result.confidence_score,
        analyzed_at=result.analyzed_at,
        geo=geo,
        removed=False,
        removal_reason=None,
    )


async def _analyze_search_result(
    result: SearchResult,
    classifier: GeminiClassifier,
    force_reanalyze: bool,
    db: Session,
) -> AnalyzeResultItem:
    if result.name:
        cleaned_name = clean_facebook_name(result.name)
        if cleaned_name and cleaned_name != result.name:
            result.name = cleaned_name

    if result.location:
        cleaned_loc = clean_facebook_location(result.location)
        if cleaned_loc and cleaned_loc != result.location:
            result.location = cleaned_loc

    if result.user_type is not None and not force_reanalyze:
        return _format_analyze_item(
            result=result,
            success=True,
            message="Skipped: already analyzed",
        )

    if not result.post_content or not result.post_content.strip():
        result.user_type = UserType.UNKNOWN
        result.confidence_score = 0.0
        result.analysis_message = "No post content available"
        result.analyzed_at = datetime.now(timezone.utc)
        result.enrichable = is_enrichable(result.name, result.location)
        return _format_analyze_item(
            result=result,
            success=True,
            message="Analyzed with fallback: no post content",
        )

    try:
        logger.info(
            "[GEO_DEBUG] analyze flow search_result id=%s location=%r name=%r",
            result.id,
            (result.location or "")[:300],
            (result.name or "")[:200],
        )
        geo_raw = await classifier.classify_geo(
            location=result.location or "",
            post_content=result.post_content or "",
            user_name=result.name or "",
        )
        logger.info("[GEO_DEBUG] analyze flow geo_raw after classify_geo: %s", geo_raw)
        geo_out = _geo_raw_to_out(geo_raw)

        if not coerce_is_us_boolean(geo_raw.get("is_us")):
            rid = result.id
            delete_search_result_and_comments(db, result)
            return AnalyzeResultItem(
                id=rid,
                success=True,
                message="Removed: non-US (geo)",
                user_type=None,
                confidence_score=None,
                analyzed_at=None,
                geo=geo_out,
                removed=True,
                removal_reason="non_us",
            )

        analysis = await classifier.classify_user(
            post_content=result.post_content,
            user_name=result.name or "",
        )

        if should_remove_not_tutoring_related(analysis):
            rid = result.id
            delete_search_result_and_comments(db, result)
            return AnalyzeResultItem(
                id=rid,
                success=True,
                message="Removed: not tutoring-related",
                user_type=None,
                confidence_score=None,
                analyzed_at=None,
                geo=geo_out,
                removed=True,
                removal_reason="not_tutoring",
            )

        user_type_map = {
            "CUSTOMER": UserType.CUSTOMER,
            "TUTOR": UserType.TUTOR,
            "UNKNOWN": UserType.UNKNOWN,
        }
        result.user_type = user_type_map.get(
            str(analysis.get("type", "UNKNOWN")).upper(),
            UserType.UNKNOWN,
        )
        result.confidence_score = max(
            0.0,
            min(1.0, float(analysis.get("confidence", 0.0))),
        )
        result.analysis_message = str(analysis.get("reason") or "")
        result.analyzed_at = datetime.now(timezone.utc)
        result.enrichable = is_enrichable(result.name, result.location)

        if result.post_date:
            parsed_ts = parse_facebook_date(result.post_date)
            if parsed_ts:
                result.post_date_timestamp = parsed_ts

        return _format_analyze_item(
            result=result,
            success=True,
            message="Analyzed successfully",
            geo=geo_out,
        )
    except Exception as exc:
        return AnalyzeResultItem(
            id=result.id,
            success=False,
            message=f"Analysis failed: {exc}",
            user_type=result.user_type.value if result.user_type else None,
            confidence_score=result.confidence_score,
            analyzed_at=result.analyzed_at,
        )


@router.get("/", response_model=SearchResultList)
async def get_results(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    status: Optional[ResultStatus] = None,
    keyword: Optional[str] = None,
    q: Optional[str] = Query(None, description="Search across name, location, and post content"),
    user_type: Optional[str] = None,
    analyzed: Optional[bool] = Query(
        None,
        description="Filter by analysis status: true (analyzed) or false (not analyzed)",
    ),
    sort_by: Optional[str] = Query(
        "post_date_timestamp",
        description="Sort field: post_date_timestamp, scraped_at, post_date, confidence_score, analyzed_at, name, status",
    ),
    sort_order: Optional[str] = Query(
        "desc",
        description="Sort order: asc or desc",
    ),
    db: Session = Depends(get_db),
):
    """Get search results with filters."""
    query = db.query(SearchResult).filter(_NOT_ARCHIVED)

    if status:
        query = query.filter(SearchResult.status == status)
    if keyword:
        query = query.filter(SearchResult.search_keyword.ilike(f"%{keyword}%"))
    if q:
        search_term = f"%{q.strip()}%"
        query = query.filter(
            SearchResult.name.ilike(search_term)
            | SearchResult.location.ilike(search_term)
            | SearchResult.post_content.ilike(search_term)
        )
    if user_type:
        query = query.filter(cast(SearchResult.user_type, String) == user_type)
    if analyzed is True:
        query = query.filter(SearchResult.analyzed_at.isnot(None))
    elif analyzed is False:
        query = query.filter(SearchResult.analyzed_at.is_(None))

    sort_map = {
        "scraped_at": SearchResult.scraped_at,
        "post_date": SearchResult.post_date,
        "post_date_timestamp": SearchResult.post_date_timestamp,
        "confidence_score": SearchResult.confidence_score,
        "analyzed_at": SearchResult.analyzed_at,
        "name": SearchResult.name,
        "status": SearchResult.status,
    }
    sort_col = sort_map.get((sort_by or "post_date_timestamp").strip().lower(), SearchResult.post_date_timestamp)
    order = (sort_order or "desc").strip().lower()
    if order == "asc":
        query = query.order_by(sort_col.asc().nullslast())
    else:
        query = query.order_by(sort_col.desc().nullslast())

    total = query.count()
    results = query.offset(skip).limit(limit).all()

    return SearchResultList(
        total=total,
        items=[SearchResultResponse.model_validate(r) for r in results],
    )


def _recent_item_from_result(r: SearchResult) -> RecentProcessedItem:
    return RecentProcessedItem(
        id=r.id,
        search_result_id=r.id,
        name=r.name or "",
        search_keyword=r.search_keyword or "",
        post_url=r.post_url,
        location=r.location,
        user_type=r.user_type.value if r.user_type else None,
        scraped_at=r.scraped_at,
        analyzed_at=r.analyzed_at,
        enriched_at=r.enriched_at,
        geo_filtered_at=r.geo_filtered_at,
        is_us=r.is_us,
    )


@router.get("/recent", response_model=RecentProcessedList)
async def get_recent_processed(
    process_type: str = Query(
        "scraped",
        description="scraped | analyzed | enriched | geo_filtered | comment_analyzed",
    ),
    skip: int = Query(0, ge=0),
    limit: int = Query(10, ge=1, le=50),
    db: Session = Depends(get_db),
):
    """Paginated last-processed entries for Jobs page."""
    process_type = (process_type or "scraped").strip().lower()

    if process_type == "scraped":
        q = (
            db.query(SearchResult)
            .filter(_NOT_ARCHIVED)
            .order_by(SearchResult.scraped_at.desc().nullslast())
        )
        total = q.count()
        rows = q.offset(skip).limit(limit).all()
        return RecentProcessedList(total=total, items=[_recent_item_from_result(r) for r in rows])

    if process_type == "analyzed":
        q = (
            db.query(SearchResult)
            .filter(_NOT_ARCHIVED, SearchResult.analyzed_at.isnot(None))
            .order_by(SearchResult.analyzed_at.desc())
        )
        total = q.count()
        rows = q.offset(skip).limit(limit).all()
        return RecentProcessedList(total=total, items=[_recent_item_from_result(r) for r in rows])

    if process_type == "enriched":
        q = (
            db.query(SearchResult)
            .filter(_NOT_ARCHIVED, SearchResult.enriched_at.isnot(None))
            .order_by(SearchResult.enriched_at.desc())
        )
        total = q.count()
        rows = q.offset(skip).limit(limit).all()
        return RecentProcessedList(total=total, items=[_recent_item_from_result(r) for r in rows])

    if process_type == "geo_filtered":
        q = (
            db.query(SearchResult)
            .filter(_NOT_ARCHIVED, SearchResult.geo_filtered_at.isnot(None))
            .order_by(SearchResult.geo_filtered_at.desc())
        )
        total = q.count()
        rows = q.offset(skip).limit(limit).all()
        return RecentProcessedList(total=total, items=[_recent_item_from_result(r) for r in rows])

    if process_type == "comment_analyzed":
        q = (
            db.query(PostComment, SearchResult)
            .join(SearchResult, PostComment.search_result_id == SearchResult.id)
            .filter(
                PostComment.archived.is_(False),
                SearchResult.archived.is_(False),
                PostComment.analyzed_at.isnot(None),
            )
            .order_by(PostComment.analyzed_at.desc())
        )
        total = q.count()
        batch = q.offset(skip).limit(limit).all()
        items: List[RecentProcessedItem] = []
        for c, sr in batch:
            text = (c.comment_text or "").strip()
            preview = text if len(text) <= 200 else text[:197] + "..."
            items.append(
                RecentProcessedItem(
                    id=c.id,
                    search_result_id=sr.id,
                    name=c.author_name or "—",
                    search_keyword=sr.search_keyword or "",
                    post_url=sr.post_url,
                    location=sr.location,
                    user_type=c.user_type.value if c.user_type else None,
                    scraped_at=c.scraped_at,
                    analyzed_at=c.analyzed_at,
                    enriched_at=None,
                    geo_filtered_at=None,
                    is_us=None,
                    lead_name=sr.name or "",
                    comment_preview=preview or None,
                )
            )
        return RecentProcessedList(total=total, items=items)

    q = (
        db.query(SearchResult)
        .filter(_NOT_ARCHIVED)
        .order_by(SearchResult.scraped_at.desc().nullslast())
    )
    total = q.count()
    rows = q.offset(skip).limit(limit).all()
    return RecentProcessedList(total=total, items=[_recent_item_from_result(r) for r in rows])


@router.get("/export/enriched")
async def export_enriched_xlsx(db: Session = Depends(get_db)):
    """Export enriched results that have at least one phone number and a location as .xlsx."""

    rows = (
        db.query(SearchResult)
        .filter(
            _NOT_ARCHIVED,
            SearchResult.enriched_at.isnot(None),
            SearchResult.enriched_phones.isnot(None),
            SearchResult.location.isnot(None),
            SearchResult.location != "",
        )
        .order_by(SearchResult.enriched_at.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Enriched Leads"

    headers = [
        "Name",
        "Location",
        "Phone Numbers",
        "Emails",
        "Addresses",
        "Age",
        "User Type",
        "Confidence",
        "Post Date",
        "Post URL",
        "Profile URL",
        "Keyword",
        "Enriched At",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    seen_names: set[str] = set()
    row_num = 2
    for r in rows:
        norm_name = (r.name or "").strip().lower()
        if not norm_name or norm_name in seen_names:
            continue
        phones = r.enriched_phones or []
        if not phones:
            continue
        phone_str = "; ".join(p.get("number", "") for p in phones if p.get("number"))
        if not phone_str:
            continue
        seen_names.add(norm_name)
        email_str = "; ".join(r.enriched_emails or [])
        addr_parts = []
        for a in r.enriched_addresses or []:
            addr_parts.append(
                ", ".join(filter(None, [a.get("street"), a.get("unit"), a.get("city"), a.get("state"), a.get("zip")]))
            )
        addr_str = "; ".join(addr_parts)

        values = [
            r.name or "",
            r.location or "",
            phone_str,
            email_str,
            addr_str,
            r.enriched_age or "",
            r.user_type.value if r.user_type else "",
            f"{r.confidence_score:.0%}" if r.confidence_score is not None else "",
            r.post_date or "",
            r.post_url or "",
            r.profile_url or "",
            r.search_keyword or "",
            r.enriched_at.isoformat() if r.enriched_at else "",
        ]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_num, column=col_idx, value=val)
        row_num += 1

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"enriched_leads_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{result_id}/comments", response_model=List[PostCommentResponse])
async def get_result_comments(result_id: str, db: Session = Depends(get_db)):
    """Get comments for a specific search result."""
    result = db.query(SearchResult).filter(SearchResult.id == result_id, _NOT_ARCHIVED).first()
    if not result:
        raise HTTPException(status_code=404, detail="Result not found")
    comments = (
        db.query(PostComment)
        .filter(
            PostComment.search_result_id == result_id,
            PostComment.archived.is_(False),
        )
        .order_by(PostComment.scraped_at.desc())
        .all()
    )
    return [PostCommentResponse.model_validate(c) for c in comments]


@router.get("/{result_id}", response_model=SearchResultResponse)
async def get_result(result_id: str, db: Session = Depends(get_db)):
    """Get a specific search result."""
    result = (
        db.query(SearchResult)
        .filter(SearchResult.id == result_id, _NOT_ARCHIVED)
        .first()
    )
    if not result:
        raise HTTPException(status_code=404, detail="Result not found")
    return result


@router.patch("/{result_id}", response_model=SearchResultResponse)
async def update_result(
    result_id: str,
    update: SearchResultUpdate,
    db: Session = Depends(get_db),
):
    """Update a search result (any editable field)."""
    result = (
        db.query(SearchResult)
        .filter(SearchResult.id == result_id, _NOT_ARCHIVED)
        .first()
    )
    if not result:
        raise HTTPException(status_code=404, detail="Result not found")

    data = update.model_dump(exclude_unset=True)
    user_type_raw = data.pop("user_type", None)
    status_raw = data.pop("status", None)
    for key, value in data.items():
        if hasattr(result, key):
            setattr(result, key, value)
    if user_type_raw is not None:
        try:
            result.user_type = UserType(user_type_raw) if user_type_raw else None
        except ValueError:
            result.user_type = None
    if status_raw is not None:
        try:
            result.status = ResultStatus(status_raw)
        except ValueError:
            pass

    result.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(result)
    return result


@router.post("/{result_id}/analyze", response_model=AnalyzeSingleResponse)
async def analyze_single_result(
    result_id: UUID,
    force_reanalyze: bool = Query(True),
    db: Session = Depends(get_db),
):
    """Analyze a single result using Gemini and persist the classification."""
    result = db.query(SearchResult).filter(SearchResult.id == result_id, _NOT_ARCHIVED).first()
    if not result:
        raise HTTPException(status_code=404, detail="Result not found")

    try:
        classifier = GeminiClassifier()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    analyzed_item = await _analyze_search_result(
        result=result,
        classifier=classifier,
        force_reanalyze=force_reanalyze,
        db=db,
    )

    if analyzed_item.success:
        db.commit()
        if analyzed_item.message.startswith("Removed:"):
            # Row deleted in DB; response already has geo / removed / removal_reason — do not overwrite.
            pass
        else:
            db.refresh(result)
            analyzed_item = _format_analyze_item(
                result=result,
                success=True,
                message=analyzed_item.message,
                geo=analyzed_item.geo,
            )
    else:
        db.rollback()

    return AnalyzeSingleResponse(item=analyzed_item)


@router.post("/analyze/batch", response_model=AnalyzeBatchResponse)
async def analyze_batch_results(
    request: AnalyzeBatchRequest,
    db: Session = Depends(get_db),
):
    """Analyze selected results in batch using Gemini."""
    if not request.result_ids:
        raise HTTPException(status_code=400, detail="result_ids is required")

    try:
        classifier = GeminiClassifier()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    result_map = {
        item.id: item
        for item in db.query(SearchResult)
        .filter(SearchResult.id.in_(request.result_ids), _NOT_ARCHIVED)
        .all()
    }

    items: List[AnalyzeResultItem] = []

    for result_id in request.result_ids:
        result = result_map.get(result_id)
        if not result:
            items.append(
                AnalyzeResultItem(
                    id=result_id,
                    success=False,
                    message="Result not found",
                )
            )
            continue

        analyzed_item = await _analyze_search_result(
            result=result,
            classifier=classifier,
            force_reanalyze=request.force_reanalyze,
            db=db,
        )
        items.append(analyzed_item)

    if any(item.success for item in items):
        db.commit()
        for item in items:
            if item.success and item.message.startswith("Removed:"):
                continue
            refreshed = result_map.get(item.id)
            if refreshed:
                db.refresh(refreshed)
    else:
        db.rollback()

    normalized_items: List[AnalyzeResultItem] = []
    for item in items:
        if item.success and item.message.startswith("Removed:"):
            normalized_items.append(item)
            continue
        result = result_map.get(item.id)
        if result:
            normalized_items.append(
                _format_analyze_item(
                    result=result,
                    success=item.success,
                    message=item.message,
                    geo=item.geo,
                )
            )
        else:
            normalized_items.append(item)

    succeeded = sum(1 for item in normalized_items if item.success and item.message != "Skipped: already analyzed")
    skipped = sum(1 for item in normalized_items if item.success and item.message == "Skipped: already analyzed")
    failed = sum(1 for item in normalized_items if not item.success)

    return AnalyzeBatchResponse(
        total=len(normalized_items),
        succeeded=succeeded,
        failed=failed,
        skipped=skipped,
        items=normalized_items,
    )


from pydantic import BaseModel

class BulkDeleteRequest(BaseModel):
    ids: List[str]


class ArchiveDuplicatesResponse(BaseModel):
    archived_results: int
    archived_comments: int
    message: str


@router.post("/archive-duplicates", response_model=ArchiveDuplicatesResponse)
async def archive_duplicate_results(db: Session = Depends(get_db)):
    """
    Mark duplicate search_results as archived (by trimmed name + normalized location).
    NULL or blank location is treated as empty string, so duplicates with no location
    still match. Keeps the earliest row per pair (by scraped_at, then id). Archives
    linked comments. Archived rows are never returned by the API.
    """
    from sqlalchemy import text

    dup_sql = """
        SELECT id FROM (
            SELECT id,
                ROW_NUMBER() OVER (
                    PARTITION BY TRIM(name), TRIM(COALESCE(location, ''))
                    ORDER BY scraped_at ASC NULLS LAST, id
                ) AS rn
            FROM search_results
            WHERE archived = false
              AND name IS NOT NULL AND TRIM(name) != ''
        ) t WHERE rn > 1
    """
    try:
        ids_rows = db.execute(text(dup_sql)).fetchall()
        dup_ids = [row[0] for row in ids_rows]
        if not dup_ids:
            return ArchiveDuplicatesResponse(
                archived_results=0,
                archived_comments=0,
                message="No duplicate name+location pairs to archive.",
            )

        n_comments = (
            db.query(PostComment)
            .filter(
                PostComment.search_result_id.in_(dup_ids),
                PostComment.archived.is_(False),
            )
            .update({PostComment.archived: True}, synchronize_session=False)
        )
        n_results = (
            db.query(SearchResult)
            .filter(SearchResult.id.in_(dup_ids))
            .update({SearchResult.archived: True}, synchronize_session=False)
        )
        db.commit()
        return ArchiveDuplicatesResponse(
            archived_results=int(n_results or 0),
            archived_comments=int(n_comments or 0),
            message=f"Archived {n_results} duplicate lead(s) and {n_comments} comment row(s).",
        )
    except Exception:
        db.rollback()
        raise


@router.post("/bulk-delete")
async def bulk_delete_results(request: BulkDeleteRequest, db: Session = Depends(get_db)):
    """Bulk delete search results."""
    # SQLite/PostgreSQL supports deleting with in_
    deleted_count = db.query(SearchResult).filter(SearchResult.id.in_(request.ids)).delete(synchronize_session=False)
    db.commit()
    return {"message": f"{deleted_count} results deleted successfully"}

@router.delete("/{result_id}")
async def delete_result(result_id: str, db: Session = Depends(get_db)):
    """Delete a search result."""
    result = (
        db.query(SearchResult)
        .filter(SearchResult.id == result_id)
        .first()
    )
    if not result:
        raise HTTPException(status_code=404, detail="Result not found")

    db.delete(result)
    db.commit()

    return {"message": "Result deleted successfully"}


# ---------------------------------------------------------------------------
# EnformionGO contact enrichment
# ---------------------------------------------------------------------------

def _format_enrich_item(
    result: SearchResult,
    success: bool,
    message: str,
) -> EnrichResultItem:
    return EnrichResultItem(
        id=result.id,
        success=success,
        message=message,
        enriched_phones=result.enriched_phones,
        enriched_emails=result.enriched_emails,
        enriched_addresses=result.enriched_addresses,
        enriched_age=result.enriched_age,
        enriched_at=result.enriched_at,
    )


async def _enrich_single(
    result: SearchResult,
    service: EnformionService,
    force: bool,
) -> EnrichResultItem:
    if result.enriched_at is not None and not force:
        return _format_enrich_item(result, True, "Skipped: already enriched")

    can, reason = EnformionService.can_enrich(result.name, result.location)
    if not can:
        return _format_enrich_item(result, False, reason)

    # Primary: Contact Enrichment
    try:
        data = await service.enrich(result.name, result.location)
    except Exception as exc:
        return _format_enrich_item(result, False, f"EnformionGO API error: {exc}")

    # Fallback: Person Search (more powerful, used when primary returns no match)
    if not data.get("matched"):
        try:
            data = await service.person_search(result.name, result.location)
        except Exception as exc:
            return _format_enrich_item(result, False, f"EnformionGO Person Search error: {exc}")

    if not data.get("matched"):
        return _format_enrich_item(result, False, "No match found in EnformionGO (tried Contact Enrichment + Person Search)")

    result.enriched_phones = data.get("phones")
    result.enriched_emails = data.get("emails")
    result.enriched_addresses = data.get("addresses")
    result.enriched_age = data.get("age")
    result.enriched_at = datetime.now(timezone.utc)

    return _format_enrich_item(result, True, "Enriched successfully")


@router.post("/{result_id}/enrich", response_model=EnrichSingleResponse)
async def enrich_single_result(
    result_id: UUID,
    force: bool = Query(False, description="Re-enrich even if already enriched"),
    db: Session = Depends(get_db),
):
    """
    Enrich a single result with contact data from EnformionGO.
    Requires both name and location; returns a warning if location is missing.
    """
    result = db.query(SearchResult).filter(SearchResult.id == result_id, _NOT_ARCHIVED).first()
    if not result:
        raise HTTPException(status_code=404, detail="Result not found")

    try:
        service = EnformionService()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    item = await _enrich_single(result, service, force)

    if item.success and item.message != "Skipped: already enriched":
        db.commit()
        db.refresh(result)
        item = _format_enrich_item(result, True, item.message)
    elif not item.success:
        db.rollback()

    return EnrichSingleResponse(item=item)


@router.post("/enrich/batch", response_model=EnrichBatchResponse)
async def enrich_batch_results(
    request: EnrichBatchRequest,
    db: Session = Depends(get_db),
):
    """
    Enrich multiple results in batch. Results without location are skipped
    with a warning message explaining why enrichment was not possible.
    """
    if not request.result_ids:
        raise HTTPException(status_code=400, detail="result_ids is required")

    try:
        service = EnformionService()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    result_map = {
        r.id: r
        for r in db.query(SearchResult)
        .filter(SearchResult.id.in_(request.result_ids), _NOT_ARCHIVED)
        .all()
    }

    items: List[EnrichResultItem] = []

    for rid in request.result_ids:
        result = result_map.get(rid)
        if not result:
            items.append(
                EnrichResultItem(id=rid, success=False, message="Result not found")
            )
            continue
        item = await _enrich_single(result, service, request.force_re_enrich)
        items.append(item)

    if any(i.success and i.message != "Skipped: already enriched" for i in items):
        db.commit()
        for i in items:
            r = result_map.get(i.id)
            if r:
                db.refresh(r)

    final: List[EnrichResultItem] = []
    for i in items:
        r = result_map.get(i.id)
        if r:
            final.append(_format_enrich_item(r, i.success, i.message))
        else:
            final.append(i)

    succeeded = sum(
        1 for i in final
        if i.success and i.message not in ("Skipped: already enriched",)
    )
    skipped = sum(1 for i in final if i.message == "Skipped: already enriched")
    failed = sum(1 for i in final if not i.success)

    return EnrichBatchResponse(
        total=len(final),
        succeeded=succeeded,
        failed=failed,
        skipped=skipped,
        items=final,
    )
